import os
import io
import json
import html as _html

import pandas as pd
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from openpyxl.styles import Font

import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth

st.set_page_config(page_title="Monitoramento de Pesquisas", layout="wide")

st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&display=swap');
:root { --eixo-vinho: #962E4D; }
html, body, [data-testid="stAppViewContainer"] { background: #F4F3EF !important; }
[data-testid="stAppViewContainer"] > section > div { background: #F4F3EF; }
.block-container, [data-testid="stMainBlockContainer"] { max-width: 1320px !important; padding: 0 2rem 3rem !important; background: #F4F3EF; }
* { box-sizing: border-box; }
body, p, span, div, label, input, select, textarea { font-family: 'Montserrat', sans-serif !important; }
[data-testid="stHeader"] { display: none; }
[data-testid="stDecoration"] { display: none; }
[data-testid="stSidebar"] { background: #F4F3EF !important; border-right: 1px solid #DADAD4 !important; }
[data-testid="stSidebar"] * { font-family: 'Montserrat', sans-serif !important; font-size: 13px !important; }
/* Streamlit usa texto com fonte Material para ícones; não sobrescreva essa fonte. */
span.material-symbols-rounded, span.material-symbols-outlined, span.material-icons,
[data-testid="stIconMaterial"], [data-testid="stIconMaterial"] * {
    font-family: 'Material Symbols Rounded', 'Material Symbols Outlined', 'Material Icons' !important;
    letter-spacing: normal !important; text-transform: none !important;
}
footer { display: none !important; }
#MainMenu { display: none !important; }
[data-testid="stButton"] > button {
    font-size: 11px !important; font-weight: 500 !important; letter-spacing: 0.1em !important;
    text-transform: uppercase !important; border-radius: 0 !important;
    border: 1px solid #962E4D !important; background: transparent !important;
    color: #962E4D !important; padding: 5px 14px !important;
    transition: background 0.15s, color 0.15s !important;
}
[data-testid="stButton"] > button:hover { background: #962E4D !important; color: #fff !important; }
[data-testid="stSidebar"] [data-testid="stButton"] > button { border: 1px solid #962E4D !important; color: #962E4D !important; }
[data-testid="stSidebar"] [data-testid="stButton"] > button:hover { background: #962E4D !important; color: #fff !important; }
[data-testid="stRadio"] > label, [data-testid="stSelectbox"] > label,
[data-testid="stMultiSelect"] > label, [data-testid="stTextInput"] > label,
[data-testid="stNumberInput"] > label, [data-testid="stTextArea"] > label {
    font-size: 11px !important; font-weight: 700 !important;
    letter-spacing: 0.1em !important; text-transform: uppercase !important; color: #767672 !important;
}
[data-testid="stTabs"] [role="tablist"] { border-bottom: 1px solid #DADAD4 !important; gap: 0 !important; }
[data-testid="stTabs"] [role="tab"] {
    font-size: 11px !important; font-weight: 500 !important; letter-spacing: 0.12em !important;
    text-transform: uppercase !important; color: #767672 !important;
    padding: 10px 20px 9px !important; border-bottom: 2px solid transparent !important;
    background: transparent !important; border-radius: 0 !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] { color: #962E4D !important; border-bottom-color: #962E4D !important; }
div[data-testid="stCodeBlock"] > pre { max-height: 520px !important; overflow: auto !important; border-radius: 0 !important; }
.ge-hero { background: #962E4D; display: flex; align-items: center; padding: 36px 48px; margin: 0 -2rem 32px -2rem; }
.ge-hero-title { font-family: 'Montserrat', sans-serif; font-size: 48px; font-weight: 800; color: #fff; line-height: 1; letter-spacing: -0.01em; }
.ge-rule { font-size: 12px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase; color: #962E4D; border-top: 1.5px solid #962E4D; padding-top: 8px; margin: 20px 0 14px; }
.at-wrap { width: 100%; background: #fff; border: 1px solid #DADAD4; overflow-x: auto; margin-top: 8px; }
.at-table { width: 100%; border-collapse: collapse; font-family: 'Montserrat', sans-serif; font-size: 12.5px; }
.at-table thead th { background: #962E4D; color: #fff; font-size: 10px; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; padding: 11px 14px; text-align: left; white-space: nowrap; }
.at-table tbody tr { border-bottom: 1px solid #DADAD4; }
.at-table tbody tr:last-child { border-bottom: none; }
.at-table tbody tr:nth-child(even) { background: #F4F3EF; }
.at-table tbody tr:hover { background: #f0e8eb; }
.at-table tbody td { padding: 9px 14px; color: #111; vertical-align: top; line-height: 1.5; max-width: 360px; word-wrap: break-word; }
.at-table td a { color: #962E4D; text-decoration: none; font-weight: 500; border-bottom: 1px solid #DADAD4; }
.at-table td a:hover { border-bottom-color: #962E4D; }
.at-meta { font-family: 'Montserrat', sans-serif; font-size: 11.5px; color: #767672; margin-bottom: 10px; }
</style>""", unsafe_allow_html=True)


# ── autenticação ──────────────────────────────────────────────────────────────

def _load_auth_config():
    if "AUTH_CONFIG" in st.secrets:
        return yaml.load(st.secrets["AUTH_CONFIG"], Loader=SafeLoader)
    if os.path.exists("config.yaml"):
        with open("config.yaml", "r", encoding="utf-8") as f:
            return yaml.load(f, Loader=SafeLoader)
    return None


_auth_cfg = _load_auth_config()
if not _auth_cfg:
    st.error("Auth não configurado. Defina AUTH_CONFIG nos Secrets ou crie config.yaml local.")
    st.stop()

authenticator = stauth.Authenticate(
    _auth_cfg["credentials"],
    _auth_cfg["cookie"]["name"],
    _auth_cfg["cookie"]["key"],
    _auth_cfg["cookie"]["expiry_days"],
)

try:
    authenticator.login(location="main")
except TypeError:
    authenticator.login("Login", "main")

authentication_status = st.session_state.get("authentication_status", None)
name  = st.session_state.get("name", "")
username = st.session_state.get("username", "")

if authentication_status is False:
    st.error("Usuário ou senha inválidos.")
    st.stop()

if authentication_status is None:
    st.info("Faça login para continuar.")
    st.stop()


# ── tema Eixo ─────────────────────────────────────────────────────────────────

EIXO = {
    "preto":         "#000000",
    "cinza":         "#999999",
    "gelo":          "#F4F3EF",
    "vinho":         "#962E4D",
    "azul":          "#192D4E",
    "amarelo":       "#E8A600",
    "amarelo_claro": "#f0d46c",
    "vermelho":      "#B84349",
}

LOGO_PATH = "Marca_eixo_vetor_Logo horizontal magenta.png"

# estilos já injetados no bloco único acima


# ── configurações ─────────────────────────────────────────────────────────────

PESQELE_SHEET_ID = (
    st.secrets.get("PESQELE_SHEET_ID")
    or os.getenv("PESQELE_SHEET_ID", "1OEmfn_RyTgrkPenzXlc6qvySs8rbVV39qmuHoULwtjQ")
).strip()

PESQELE_ABA_GID = int(
    (st.secrets.get("PESQELE_ABA_GID") or os.getenv("PESQELE_ABA_GID", "1376656858")).strip()
)


# ── Google Sheets ─────────────────────────────────────────────────────────────

@st.cache_resource
def _gs_client():
    if "GOOGLE_SHEETS_CREDS" not in st.secrets:
        raise RuntimeError("GOOGLE_SHEETS_CREDS não configurado nos Secrets.")
    raw = st.secrets["GOOGLE_SHEETS_CREDS"]
    creds_dict = json.loads(raw) if isinstance(raw, str) else dict(raw)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.authorize(credentials)


@st.cache_data(ttl=300)
def load_pesqele(spreadsheet_id: str, gid: int, cache_bust: int = 0) -> pd.DataFrame:
    gc = _gs_client()
    sh = gc.open_by_key(spreadsheet_id)

    ws = None
    for w in sh.worksheets():
        if int(getattr(w, "id", -1)) == int(gid):
            ws = w
            break
    if ws is None:
        raise RuntimeError(f"Aba com gid={gid} não encontrada.")

    values = ws.get_all_values()
    if not values:
        return pd.DataFrame()

    df = pd.DataFrame(values[1:], columns=values[0])

    for col in df.columns:
        cn = col.strip().lower()
        if cn in ("data_registro", "data_divulgacao", "data"):
            df[col] = pd.to_datetime(df[col], errors="coerce")

    return df


def to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Filtrado")
        ws = writer.sheets["Filtrado"]
        ws.freeze_panes = "A2"

        link_col = None
        for i, col in enumerate(df.columns, start=1):
            if str(col).strip().lower() in ("link", "url", "href"):
                link_col = i
                break
        if link_col:
            hf = Font(color="0563C1", underline="single")
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=link_col)
                url = str(cell.value or "").strip()
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = hf

        for col_idx, col_name in enumerate(df.columns, start=1):
            try:
                max_len = max(
                    [len(str(col_name))]
                    + [len(str(v)) for v in df[col_name].astype(str).head(200)]
                )
            except Exception:
                max_len = len(str(col_name))
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = min(max(10, max_len + 2), 60)

    buf.seek(0)
    return buf.read()


# ── componente de filtro multiselect com popover ──────────────────────────────

def dropdown_multiselect(label: str, options: list, default: list, key: str) -> list:
    def _ensure(k, d):
        if k not in st.session_state:
            st.session_state[k] = d

    _ensure(key, list(default))
    options = sorted([o for o in options if str(o).strip()], key=str)
    sel = st.session_state[key]
    n, total = len(sel), len(options)
    badge = "Todos" if n == total else ("0 selecionados" if n == 0 else f"{n} selecionados")

    with st.popover(f"{label}   —   {badge}", use_container_width=True):
        qk = f"{key}__q"
        _ensure(qk, "")
        q = st.text_input("Pesquisar", key=qk).strip().lower()

        c1, c2 = st.columns(2)

        def _ck(o):
            return f"{key}__opt__{hash(str(o))}"

        if c1.button("Selecionar tudo", key=f"{key}__all", use_container_width=True):
            st.session_state[key] = list(options)
            for o in options:
                st.session_state[_ck(o)] = True
            st.rerun()

        if c2.button("Limpar", key=f"{key}__none", use_container_width=True):
            st.session_state[key] = []
            for o in options:
                st.session_state[_ck(o)] = False
            st.rerun()

        st.markdown("---")
        filtered = [o for o in options if q in str(o).lower()] if q else options

        for o in filtered[:400]:
            ck = _ck(o)
            _ensure(ck, o in st.session_state[key])
            checked = st.checkbox(str(o), key=ck)
            changed = False
            if checked and o not in st.session_state[key]:
                st.session_state[key].append(o)
                changed = True
            if not checked and o in st.session_state[key]:
                st.session_state[key].remove(o)
                changed = True
            if changed:
                st.rerun()

        if len(filtered) > 400:
            st.caption(f"Mostrando 400 de {len(filtered)} (use a busca).")

    return list(st.session_state[key])


# ── session state ─────────────────────────────────────────────────────────────

if "cache_bust_pesqele" not in st.session_state:
    st.session_state.cache_bust_pesqele = 0


# ── carrega dados ─────────────────────────────────────────────────────────────

try:
    df = load_pesqele(PESQELE_SHEET_ID, PESQELE_ABA_GID, st.session_state.cache_bust_pesqele)
except Exception as e:
    st.error(f"Erro ao carregar a planilha: {e}")
    st.stop()

if df.empty:
    st.warning("Planilha vazia ou sem cabeçalho.")
    st.stop()


# ── detecta colunas dinamicamente ─────────────────────────────────────────────

def _find_col(df, *candidates):
    for c in df.columns:
        if c.strip().lower() in [x.lower() for x in candidates]:
            return c
    return None

col_abrangencia  = _find_col(df, "abrangencia", "abrangência")
col_empresa      = _find_col(df, "empresa_contratada", "empresa contratada", "empresa")
col_cargos       = _find_col(df, "cargos", "cargo")
col_data_reg     = _find_col(df, "data_registro", "data registro", "data_reg")
col_data_div     = _find_col(df, "data_divulgacao", "data_divulgação", "data divulgacao", "data divulgação")


# ── render HTML table ─────────────────────────────────────────────────────────

_TABLE_URL_COLS = {"link", "url", "href", "link direto"}

def _render_table_html(df: pd.DataFrame) -> str:
    headers = "".join(f'<th>{_html.escape(str(c))}</th>' for c in df.columns)
    rows = ""
    for _, row in df.iterrows():
        cells = ""
        for col_name, val in row.items():
            raw = str(val).strip() if pd.notna(val) else ""
            if not raw or raw.lower() in ("nan", "none"):
                cells += "<td></td>"
            elif col_name.strip().lower() in _TABLE_URL_COLS or raw.startswith("http://") or raw.startswith("https://"):
                safe = _html.escape(raw)
                cells += f'<td><a href="{safe}" target="_blank" rel="noopener">↗ link</a></td>'
            else:
                safe = _html.escape(raw)
                if len(safe) > 200:
                    cells += f'<td><span title="{safe}" style="cursor:help;">{safe[:200]}…</span></td>'
                else:
                    cells += f'<td>{safe}</td>'
        rows += f"<tr>{cells}</tr>"
    return (
        f'<div class="at-wrap"><table class="at-table">'
        f'<thead><tr>{headers}</tr></thead>'
        f'<tbody>{rows}</tbody>'
        f'</table></div>'
    )


# ── sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    try:
        st.image(LOGO_PATH, use_container_width=True)
    except Exception:
        st.caption("Logo não encontrada.")
    st.markdown(
        '<div style="border-left:3px solid #962E4D;padding:10px 12px;'
        'margin:10px 0 0 0;background:#fff;border-radius:0 4px 4px 0;">'
        '<p style="font-family:Montserrat,sans-serif;font-size:12.5px;'
        'color:#111;line-height:1.65;margin:0;">'
        'Consolidado de <strong>pesquisas eleitorais</strong> registradas no TSE.'
        '</p></div>',
        unsafe_allow_html=True,
    )
    st.markdown("---")
    if st.button("↻ Recarregar dados", use_container_width=True, key="btn_reload_sidebar_pesq"):
        st.cache_data.clear()
        st.rerun()
    st.markdown("---")
    st.caption(f"Usuário: **{st.session_state.get('name', '')}** ({st.session_state.get('username', '')})")
    if _auth_cfg:
        authenticator.logout("Sair", "sidebar")


# ── página principal ──────────────────────────────────────────────────────────

st.markdown(
    '<div class="ge-hero"><div class="ge-hero-title">Monitoramento de Pesquisas</div></div>',
    unsafe_allow_html=True,
)

# ── filtros ───────────────────────────────────────────────────────────────────

selected = {}
sel_cargos = []
sel_reg = None
sel_div = None

col_f1, col_f2, col_f3 = st.columns(3)
with col_f1:
    if col_abrangencia:
        opts = df[col_abrangencia].dropna().unique().tolist()
        selected[col_abrangencia] = dropdown_multiselect(
            "Abrangência", opts, opts, key="flt_abrangencia"
        )

with col_f2:
    if col_empresa:
        opts = df[col_empresa].dropna().unique().tolist()
        selected[col_empresa] = dropdown_multiselect(
            "Empresa contratada", opts, opts, key="flt_empresa"
        )

with col_f3:
    if col_cargos:
        todos_cargos = set()
        for val in df[col_cargos].dropna():
            for c in str(val).split(","):
                c = c.strip()
                if c:
                    todos_cargos.add(c)
        opts_cargos = sorted(todos_cargos)
        sel_cargos = dropdown_multiselect(
            "Cargos", opts_cargos, opts_cargos, key="flt_cargos"
        )

col_f4, col_f5, col_f6 = st.columns(3)
with col_f4:
    if col_data_reg and df[col_data_reg].notna().any():
        d0 = df[col_data_reg].min()
        d1 = df[col_data_reg].max()
        sel_reg = st.date_input(
            "Data de registro",
            value=(d0.date(), d1.date()),
            min_value=d0.date(),
            max_value=d1.date(),
            key="flt_data_reg",
            format="DD/MM/YYYY",
        )

with col_f5:
    if col_data_div and df[col_data_div].notna().any():
        d0 = df[col_data_div].min()
        d1 = df[col_data_div].max()
        sel_div = st.date_input(
            "Data de divulgação",
            value=(d0.date(), d1.date()),
            min_value=d0.date(),
            max_value=d1.date(),
            key="flt_data_div",
            format="DD/MM/YYYY",
        )

with col_f6:
    q_busca = st.text_input(
        "Filtrar", placeholder="Buscar por palavra-chave…", key="q_busca_pesqele"
    ).strip().lower()

# ── aplica filtros ────────────────────────────────────────────────────────────

f = df.copy()

for col, vals in selected.items():
    if vals is not None:
        if len(vals) == 0:
            f = f.iloc[0:0]
            break
        f = f[f[col].isin(vals)]

if col_cargos and sel_cargos and len(sel_cargos) < len(sorted(
    {c.strip() for v in df[col_cargos].dropna() for c in str(v).split(",") if c.strip()}
)):
    mask = f[col_cargos].apply(
        lambda val: any(
            c.strip() in sel_cargos
            for c in str(val).split(",")
        ) if pd.notna(val) else False
    )
    f = f[mask]

if col_data_reg and sel_reg and len(sel_reg) == 2:
    a, b = sel_reg
    f = f[
        (f[col_data_reg] >= pd.Timestamp(a)) &
        (f[col_data_reg] <= pd.Timestamp(b))
    ]

if col_data_div and sel_div and len(sel_div) == 2:
    a, b = sel_div
    f = f[
        (f[col_data_div] >= pd.Timestamp(a)) &
        (f[col_data_div] <= pd.Timestamp(b))
    ]

if q_busca and not f.empty:
    mask = pd.Series(False, index=f.index)
    for col in f.columns:
        mask = mask | f[col].astype(str).str.lower().str.contains(q_busca, na=False)
    f = f[mask]

COLUNAS_OCULTAS = {"eleicao", "eleição", "uf_filtro", "capturado em", "capturado_em"}
colunas_visiveis = [c for c in f.columns if c.strip().lower() not in COLUNAS_OCULTAS]
f_vis = f[colunas_visiveis]

st.markdown('<div class="ge-rule"></div>', unsafe_allow_html=True)

col_meta, col_dl = st.columns([3, 1])
with col_meta:
    n = len(f_vis)
    st.markdown(
        f'<p class="at-meta">{n} registro{"s" if n != 1 else ""} encontrado{"s" if n != 1 else ""} '
        f'({len(df)} no total)</p>',
        unsafe_allow_html=True,
    )
with col_dl:
    st.download_button(
        label=f"⬇ Baixar seleção ({len(f_vis)})",
        data=to_xlsx_bytes(f_vis),
        file_name="pesquisas_filtradas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_pesqele",
    )

st.markdown(_render_table_html(f_vis), unsafe_allow_html=True)
