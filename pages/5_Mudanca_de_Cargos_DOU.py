from datetime import datetime
import os
import io
import re
import json
import html as _html

import pandas as pd
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from openpyxl.styles import Font

import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth


st.set_page_config(page_title="Mudança de cargos no DOU", layout="wide")

st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&display=swap');
:root { --eixo-vinho: #962E4D; }
html, body, [data-testid="stAppViewContainer"] { background: #F4F3EF !important; }
[data-testid="stAppViewContainer"] > section > div { background: #F4F3EF; }
.block-container, [data-testid="stMainBlockContainer"] { max-width: 1320px !important; padding: 0 2rem 3rem !important; background: #F4F3EF; }
* { box-sizing: border-box; }
body, p, span, div, label, input, select, textarea { font-family: 'Montserrat', sans-serif !important; }
[data-testid="stHeader"] { display: none; }
[data-testid="stDecoration"] { display: none; }
[data-testid="stSidebar"] { background: #F4F3EF !important; border-right: 1px solid #DADAD4 !important; }
[data-testid="stSidebar"] * { font-family: 'Montserrat', sans-serif !important; font-size: 13px !important; }
/* Streamlit usa texto com fonte Material para ícones; não sobrescreva essa fonte. */
span.material-symbols-rounded, span.material-symbols-outlined, span.material-icons,
[data-testid="stIconMaterial"], [data-testid="stIconMaterial"] * {
    font-family: 'Material Symbols Rounded', 'Material Symbols Outlined', 'Material Icons' !important;
    letter-spacing: normal !important; text-transform: none !important;
}
footer { display: none !important; }
#MainMenu { display: none !important; }
[data-testid="stButton"] > button {
    font-size: 11px !important; font-weight: 500 !important; letter-spacing: 0.1em !important;
    text-transform: uppercase !important; border-radius: 0 !important;
    border: 1px solid #962E4D !important; background: transparent !important;
    color: #962E4D !important; padding: 5px 14px !important;
    transition: background 0.15s, color 0.15s !important;
}
[data-testid="stButton"] > button:hover { background: #962E4D !important; color: #fff !important; }
[data-testid="stSidebar"] [data-testid="stButton"] > button { border: 1px solid #962E4D !important; color: #962E4D !important; }
[data-testid="stSidebar"] [data-testid="stButton"] > button:hover { background: #962E4D !important; color: #fff !important; }
[data-testid="stRadio"] > label, [data-testid="stSelectbox"] > label,
[data-testid="stMultiSelect"] > label, [data-testid="stTextInput"] > label,
[data-testid="stNumberInput"] > label, [data-testid="stTextArea"] > label {
    font-size: 11px !important; font-weight: 700 !important;
    letter-spacing: 0.1em !important; text-transform: uppercase !important; color: #767672 !important;
}
[data-testid="stTabs"] [role="tablist"] { border-bottom: 1px solid #DADAD4 !important; gap: 0 !important; }
[data-testid="stTabs"] [role="tab"] {
    font-size: 11px !important; font-weight: 500 !important; letter-spacing: 0.12em !important;
    text-transform: uppercase !important; color: #767672 !important;
    padding: 10px 20px 9px !important; border-bottom: 2px solid transparent !important;
    background: transparent !important; border-radius: 0 !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] { color: #962E4D !important; border-bottom-color: #962E4D !important; }
.ge-hero { background: #962E4D; display: flex; align-items: center; padding: 36px 48px; margin: 0 -2rem 32px -2rem; }
.ge-hero-title { font-family: 'Montserrat', sans-serif; font-size: 48px; font-weight: 800; color: #fff; line-height: 1; letter-spacing: -0.01em; }
.ge-rule { font-size: 12px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase; color: #962E4D; border-top: 1.5px solid #962E4D; padding-top: 8px; margin: 20px 0 14px; }
.at-wrap { width: 100%; background: #fff; border: 1px solid #DADAD4; overflow-x: auto; margin-top: 8px; }
.at-table { width: 100%; border-collapse: collapse; font-family: 'Montserrat', sans-serif; font-size: 12.5px; }
.at-table thead th { background: #962E4D; color: #fff; font-size: 10px; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; padding: 11px 14px; text-align: left; white-space: nowrap; }
.at-table tbody tr { border-bottom: 1px solid #DADAD4; }
.at-table tbody tr:last-child { border-bottom: none; }
.at-table tbody tr:nth-child(even) { background: #F4F3EF; }
.at-table tbody tr:hover { background: #f0e8eb; }
.at-table tbody td { padding: 9px 14px; color: #111; vertical-align: top; line-height: 1.5; max-width: 360px; word-wrap: break-word; }
.at-table td a { color: #962E4D; text-decoration: none; font-weight: 500; border-bottom: 1px solid #DADAD4; }
.at-table td a:hover { border-bottom-color: #962E4D; }
.at-meta { font-family: 'Montserrat', sans-serif; font-size: 11.5px; color: #767672; margin-bottom: 10px; }
</style>""", unsafe_allow_html=True)


def _load_auth_config():
    if "AUTH_CONFIG" in st.secrets:
        return yaml.load(st.secrets["AUTH_CONFIG"], Loader=SafeLoader)

    if os.path.exists("config.yaml"):
        with open("config.yaml", "r", encoding="utf-8") as f:
            return yaml.load(f, Loader=SafeLoader)

    return None


_auth_cfg = _load_auth_config()
if not _auth_cfg:
    st.error("Auth não configurado. Defina AUTH_CONFIG nos Secrets ou crie config.yaml local.")
    st.stop()

authenticator = stauth.Authenticate(
    _auth_cfg["credentials"],
    _auth_cfg["cookie"]["name"],
    _auth_cfg["cookie"]["key"],
    _auth_cfg["cookie"]["expiry_days"],
)

# O campo branco do login vem daqui, e tem que ser desenhado antes do gate:
# sem ninguém logado o gate chama st.stop() e nada depois dele roda.
from ui_login import CSS_LOGIN
st.markdown(CSS_LOGIN, unsafe_allow_html=True)

try:
    authenticator.login(location="main")
except TypeError:
    authenticator.login("Login", "main")

authentication_status = st.session_state.get("authentication_status", None)
name = st.session_state.get("name", "")
username = st.session_state.get("username", "")

if authentication_status is False:
    st.error("Usuário ou senha inválidos.")
    st.stop()

if authentication_status is None:
    st.info("Faça login para continuar.")
    st.stop()


# ========= TEMA EIXO =========
EIXO = {
    "preto": "#000000",
    "cinza": "#999999",
    "gelo":          "#F4F3EF",
    "vinho": "#962E4D",
    "azul": "#192D4E",
    "amarelo": "#E8A600",
    "amarelo_claro": "#f0d46c",
    "vermelho": "#B84349",
}

LOGO_PATH = "Marca_eixo_vetor_Logo horizontal magenta.png"

# estilos já injetados no bloco único acima


# ========= CONFIG (Secrets) =========
PLANILHA_ID = (st.secrets.get("CARGOS_SHEET_ID") or os.getenv("CARGOS_SHEET_ID", "")).strip()
ABA_GID = int((st.secrets.get("CARGOS_ABA_GID") or os.getenv("CARGOS_ABA_GID", "0")).strip() or 0)


# ========= GOOGLE SHEETS CLIENT =========
@st.cache_resource
def _gs_client():
    if "GOOGLE_SHEETS_CREDS" not in st.secrets:
        raise RuntimeError("GOOGLE_SHEETS_CREDS não configurado nos Secrets.")

    raw = st.secrets["GOOGLE_SHEETS_CREDS"]

    if isinstance(raw, str):
        creds_dict = json.loads(raw)
    else:
        creds_dict = dict(raw)

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.authorize(credentials)


@st.cache_data(ttl=300)
def load_sheet_df_by_gid(spreadsheet_id: str, gid: int, cache_bust: int = 0) -> pd.DataFrame:
    gc = _gs_client()
    sh = gc.open_by_key(spreadsheet_id)

    ws = None
    for w in sh.worksheets():
        if int(getattr(w, "id", -1)) == int(gid):
            ws = w
            break
    if ws is None:
        raise RuntimeError(f"Não achei worksheet com gid={gid}. Confere o gid e se a conta tem acesso.")

    values = ws.get_all_values()
    if not values:
        return pd.DataFrame()

    header = values[0]
    rows = values[1:] if len(values) > 1 else []
    df = pd.DataFrame(rows, columns=header)

    for c in df.columns:
        cn = c.strip().lower()
        if cn in ("data", "dt_publicacao", "dt", "date"):
            df[c] = pd.to_datetime(df[c], errors="coerce")
        if cn in ("valor", "value", "qtde", "qtd", "quantidade"):
            df[c] = pd.to_numeric(df[c], errors="coerce")

    return df


def to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Filtrado")
        ws = writer.sheets["Filtrado"]

        ws.freeze_panes = "A2"

        link_col = None
        for i, col in enumerate(df.columns, start=1):
            if str(col).strip().lower() in ("link", "url", "href"):
                link_col = i
                break

        if link_col is not None:
            hyperlink_font = Font(color="0563C1", underline="single")
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=link_col)
                url = str(cell.value).strip() if cell.value is not None else ""
                if url.startswith("http://") or url.startswith("https://"):
                    cell.hyperlink = url
                    cell.font = hyperlink_font

        for col_idx, col_name in enumerate(df.columns, start=1):
            try:
                max_len = max(
                    [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).head(200).tolist()]
                )
            except Exception:
                max_len = len(str(col_name))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(10, max_len + 2), 60)

    buf.seek(0)
    return buf.read()


def dropdown_multiselect(label: str, options: list, default: list, key: str):
    def _ensure_state(k: str, d):
        if k not in st.session_state:
            st.session_state[k] = d

    _ensure_state(key, list(default))

    options = [o for o in options if str(o).strip() != ""]
    options = sorted(options, key=lambda x: str(x))

    selected = st.session_state[key]
    sel_count = len(selected)
    total = len(options)

    if sel_count == 0:
        right = "0 selecionados"
    elif sel_count == total:
        right = "Todos"
    else:
        right = f"{sel_count} selecionados"

    with st.popover(f"{label}   —   {right}", use_container_width=True):
        qk = f"{key}__q"
        _ensure_state(qk, "")
        qval = st.text_input("Digite para pesquisar", key=qk).strip().lower()

        c1, c2 = st.columns(2)

        def _opt_ck(opt):
            return f"{key}__opt__{hash(str(opt))}"

        if c1.button("Selecionar tudo", key=f"{key}__all", use_container_width=True):
            st.session_state[key] = list(options)
            for o in options:
                st.session_state[_opt_ck(o)] = True
            st.rerun()

        if c2.button("Limpar", key=f"{key}__none", use_container_width=True):
            st.session_state[key] = []
            for o in options:
                st.session_state[_opt_ck(o)] = False
            st.rerun()

        st.markdown("---")

        filtered = options
        if qval:
            filtered = [o for o in options if qval in str(o).lower()]

        for o in filtered[:400]:
            ck = _opt_ck(o)
            _ensure_state(ck, o in st.session_state[key])
            new_checked = st.checkbox(str(o), key=ck)

            changed = False
            if new_checked and (o not in st.session_state[key]):
                st.session_state[key].append(o)
                changed = True
            if (not new_checked) and (o in st.session_state[key]):
                st.session_state[key].remove(o)
                changed = True

            if changed:
                st.rerun()

        if len(filtered) > 400:
            st.caption(f"Mostrando 400 de {len(filtered)} resultados (use a busca).")

    return list(st.session_state[key])


# ========= STATE =========
if "cache_bust_cargos" not in st.session_state:
    st.session_state.cache_bust_cargos = 0


# ========= LOAD DF =========
if not PLANILHA_ID:
    st.error("Faltou CARGOS_SHEET_ID nos Secrets/env.")
    st.stop()

df = load_sheet_df_by_gid(PLANILHA_ID, ABA_GID, st.session_state.cache_bust_cargos)

if df.empty:
    st.warning("Aba vazia ou sem cabeçalho.")
    st.stop()


# ========= COLS =========
date_col = next((c for c in df.columns if c.strip().lower() in ("data", "dt_publicacao", "date")), None)
col_verbo = next((c for c in df.columns if c.strip().lower() == "verbos acionados".lower()), None)
col_termo = next((c for c in df.columns if c.strip().lower() == "termos de cargo".lower()), None)
col_secao = next((c for c in df.columns if c.strip().lower() in ("seção", "secao")), None)


# ========= RENDER HTML TABLE =========
_TABLE_URL_COLS = {"link", "url", "href", "link direto"}

def _render_table_html(df: pd.DataFrame) -> str:
    headers = "".join(f'<th>{_html.escape(str(c))}</th>' for c in df.columns)
    rows = ""
    for _, row in df.iterrows():
        cells = ""
        for col_name, val in row.items():
            raw = str(val).strip() if pd.notna(val) else ""
            if not raw or raw.lower() in ("nan", "none"):
                cells += "<td></td>"
            elif col_name.strip().lower() in _TABLE_URL_COLS or raw.startswith("http://") or raw.startswith("https://"):
                safe = _html.escape(raw)
                cells += f'<td><a href="{safe}" target="_blank" rel="noopener">↗ link</a></td>'
            else:
                safe = _html.escape(raw)
                if len(safe) > 200:
                    cells += f'<td><span title="{safe}" style="cursor:help;">{safe[:200]}…</span></td>'
                else:
                    cells += f'<td>{safe}</td>'
        rows += f"<tr>{cells}</tr>"
    return (
        f'<div class="at-wrap"><table class="at-table">'
        f'<thead><tr>{headers}</tr></thead>'
        f'<tbody>{rows}</tbody>'
        f'</table></div>'
    )


# ========= SIDEBAR =========
with st.sidebar:
    try:
        st.image(LOGO_PATH, use_container_width=True)
    except Exception:
        st.caption("Logo não encontrada.")
    st.markdown(
        '<div style="border-left:3px solid #962E4D;padding:10px 12px;'
        'margin:10px 0 0 0;background:transparent;">'
        '<p style="font-family:Montserrat,sans-serif;font-size:12.5px;'
        'color:#111;line-height:1.65;margin:0;">'
        'Clipping de <strong>atos de movimentação de cargos</strong> no DOU '
        '(nomear, dispensar, exonerar, designar etc.).'
        '</p></div>',
        unsafe_allow_html=True,
    )
    st.markdown("---")
    if st.button("↻ Recarregar dados", use_container_width=True, key="btn_reload_sidebar_dou"):
        st.cache_data.clear()
        st.rerun()
    st.markdown("---")
    st.caption(f"Usuário: **{st.session_state.get('name', '')}** ({st.session_state.get('username', '')})")
    if _auth_cfg:
        authenticator.logout("Sair", "sidebar")


# ========= MAIN =========
st.markdown(
    '<div class="ge-hero"><div class="ge-hero-title">Mudança de Cargos no DOU</div></div>',
    unsafe_allow_html=True,
)

# ── filtros ───────────────────────────────────────────────────────────────────

sel_range = None
selected = {}

col_f1, col_f2, col_f3, col_f4 = st.columns(4)
with col_f1:
    if date_col and df[date_col].notna().any():
        d0 = df[date_col].min()
        d1 = df[date_col].max()
        sel_range = st.date_input(
            "Período",
            value=(d0.date(), d1.date()),
            min_value=d0.date(),
            max_value=d1.date(),
            key="flt_periodo_cargos",
            format="DD/MM/YYYY",
        )

with col_f2:
    if col_verbo:
        opts = df[col_verbo].dropna().unique().tolist()
        selected[col_verbo] = dropdown_multiselect("Verbos acionados", opts, opts, key="flt_verbo_cargos")

with col_f3:
    if col_termo:
        opts = df[col_termo].dropna().unique().tolist()
        selected[col_termo] = dropdown_multiselect("Termos de cargo", opts, opts, key="flt_termo_cargos")

with col_f4:
    if col_secao:
        opts = df[col_secao].dropna().unique().tolist()
        selected[col_secao] = dropdown_multiselect("Seção", opts, opts, key="flt_secao_cargos")

q = st.text_input(
    "Filtrar", placeholder="Buscar por palavra-chave…", key="q_busca_cargos"
).strip().lower()

# ── aplica filtros ────────────────────────────────────────────────────────────

f = df.copy()

for c, vals in selected.items():
    if vals is not None:
        if len(vals) == 0:
            f = f.iloc[0:0]
            break
        f = f[f[c].isin(vals)]

if date_col and sel_range and len(sel_range) == 2:
    a, b = sel_range
    f = f[(f[date_col] >= pd.Timestamp(a)) & (f[date_col] <= pd.Timestamp(b))]

if q and not f.empty:
    mask = False
    for c in f.columns:
        mask = mask | f[c].astype(str).str.lower().str.contains(q, na=False)
    f = f[mask]

st.markdown('<div class="ge-rule"></div>', unsafe_allow_html=True)

col_meta, col_dl = st.columns([3, 1])
with col_meta:
    n = len(f)
    st.markdown(
        f'<p class="at-meta">{n} registro{"s" if n != 1 else ""} encontrado{"s" if n != 1 else ""} '
        f'({len(df)} no total)</p>',
        unsafe_allow_html=True,
    )
with col_dl:
    st.download_button(
        label=f"⬇ Baixar seleção ({len(f)})",
        data=to_xlsx_bytes(f),
        file_name="mudanca_cargos_dou_filtrado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_cargos_dou",
    )

st.markdown(_render_table_html(f), unsafe_allow_html=True)
